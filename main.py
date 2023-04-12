from tkinter import END, filedialog, LEFT, messagebox
import openpyxl
import time
import cv2 as cv
import pyautogui
import os
import json
import tkinter
import sys
import pyperclip
from numpy import double



class GlobalData:
    DEFAULTS = {
        'start_num': 1,  # 起始位置
        'messagelist_para': 85,  # 联系人列表
        'contacts_para': 85,  # 联系人&群偏移
        'text_para': 100,  # 文本框偏移
        'search_wait_time': 2.0,  # 搜索联系人等待时间
        'message_wait_time': 1.5,  # 两次消息间隔时间
        'path': ".\\image\\ceshi.xlsx"
    }

    def __init__(self, filepath='global_data.json'):
        self.filepath = os.path.abspath(filepath)
        self.data = self.load(filepath)

    @classmethod
    def load(cls, filepath):
        try:
            with open(filepath, 'r') as fp:
                data = json.load(fp)
        except FileNotFoundError:
            print("未找到 global_data.json 文件")
            return cls.DEFAULTS
        except json.JSONDecodeError:
            print(f"无法解析 {os.path.abspath(filepath)} 文件中的 JSON 数据")
            return cls.DEFAULTS

        missing_keys = set(cls.DEFAULTS.keys()) - set(data.keys())
        if missing_keys:
            print(f"注意：{os.path.abspath(filepath)} 文件缺少以下默认键：{list(missing_keys)}")
            data.update({key: val for key, val in cls.DEFAULTS.items() if key not in data})

        return data

    def save(self):
        with open(self.filepath, 'w') as fp:
            json.dump(self.data, fp)

    def update_data(self, new_data):
        for k, v in new_data.items():
            self.data[k] = v
        self.save()


def set_path(filepath):
    global path
    path = filepath
    pass


class Message:
    def __init__(self, global_data):
        self.global_data = global_data
        pass

    def message_start(self):

        obj_path = self.global_data.data['path']
        data = read_excel_file(obj_path)

        for i, row in enumerate(data):
            temp_image = ".\\image\\leftmenu.png"
            x, y = find_image_location(temp_image)
            click_below(x, y, self.global_data.data['messagelist_para'], 0.2)

            # 调用开始行数
            if i < self.global_data.data['start_num'] - 1:
                continue

            # 如果第一个单元格为空，跳过当前行，否则查找是否有搜索栏删除图标，如果没有就查找空白搜索栏，并输入联系人
            if check_none(row[0], self.global_data.data['search_wait_time']):
                continue
                # 判断有无联系人或者群
            if row[1] == "联系人":
                object_image = ".\\image\\friend.png"
            elif row[1] == "群聊":
                object_image = ".\\image\\group.png"
            else:
                continue
            result = check_image_existence(object_image)

            # 联系人or群不存在，直接进入下一次循环
            if result is None:
                print("第" + str(i+1) + "联系人未添加")
                continue
            # 联系人or群存在，按照给定参数点击
            else:
                print("第" + str(i+1) + "联系人存在，准备发送反馈")
                x, y = result
                click_below(x, y, self.global_data.data['contacts_para'], 0.3)#等待0.3s聊天窗口显示

            for j, cell in enumerate(row[2:]):
                cell_type = get_cell_type(cell)
                process_cell_value(cell_type, cell, self.global_data.data['message_wait_time'],
                                   self.global_data.data['text_para'])

class Group:
    def __init__(self,global_data):
        self.global_data = global_data
        pass

    def group_start(self):

        obj_path = self.global_data.data['path']
        data = read_excel_file(obj_path)

        sum = 0

        for i, row in enumerate(data):
            temp_image = ".\\image\\leftmenu.png"
            x, y = find_image_location(temp_image)
            click_below(x, y, self.global_data.data['messagelist_para'], 0.2)

            if check_none(row[0], self.global_data.data['search_wait_time']):
                continue  # 如果第一个单元格为空，跳过当前行

            # 判断有无命名群聊
            object_image = ".\\image\\group.png"
            result = check_image_existence(object_image)

            if result is None:
                object_image = ".\\image\\new_group.png"
                x, y = find_image_location(object_image)
                click_center(x, y, 0.5)
            else:
                x, y = result
                click_below(x, y, self.global_data.data['contacts_para'], 0.3)# 等待0.3s聊天框显示
                object_image = ".\\image\\add_grouper.png"
                x, y = find_image_location(object_image)
                click_center(x, y, 0.5)# 等待0.5s添加群成员窗口出现

            for j, cell in enumerate(row[1:]):
                if cell == "":
                    continue
                pyperclip.copy(cell)
                pyperclip.copy(cell)
                time.sleep(0.1)
                pyautogui.hotkey("ctrl", "v")

                time.sleep(0.5)  # 等待能否搜到群友

                object_image = ".\\image\\allowed_added_grouper.png"
                grouper = check_image_existence(object_image)
                if grouper is None:
                    object_image = ".\\image\\search_grouper_delete.png"
                    x, y = find_image_location(object_image)
                    click_center(x, y, 0.3)
                    pyautogui.move(0,-100,0.2)
                    continue
                else:
                    x, y = grouper
                    click_center(x, y, 0.3)
                    pyautogui.move(-100, 0, 0.2)
                    sum++1

            if result is None:
                object_image = ".\\image\\creat_group.png"
                x, y = find_image_location(object_image)
                click_center(x, y, 6)# 等待创建群6s

                object_image = ".\\image\\set_group.png"
                x, y = find_image_location(object_image)
                click_center(x, y, 0.2)

                object_image = ".\\image\\group_name.png"
                x, y = find_image_location(object_image)
                click_center(x, y, 0.2)

                pyperclip.copy(row[0])
                pyperclip.copy(row[0])
                pyautogui.hotkey("ctrl", "v")
                time.sleep(0.2)

                pyautogui.press("enter")
                time.sleep(1)# 修改群名后等待1s

                object_image = ".\\image\\group_remarks.png"
                x, y = find_image_location(object_image)
                click_center(x, y, 0.2)

                pyperclip.copy(row[0])
                pyperclip.copy(row[0])
                pyautogui.hotkey("ctrl", "v")
                time.sleep(0.2)

                object_image = ".\\image\\enter_change_group_remarks.png"
                x, y = find_image_location(object_image)
                click_center(x, y, 2)
            else:
                if sum==0:
                    pyautogui.press('esc')
                    continue
                object_image = ".\\image\\enter_add_grouper.png"
                x, y = find_image_location(object_image)
                click_center(x, y, 6)  # 等待创建群6s
                sum=0

class StdoutRedirector(object):
    # 重定向输出类
    def __init__(self, text_widget):
        self.text_space = text_widget
        # 将其备份
        self.stdoutbak = sys.stdout
        self.stderrbak = sys.stderr

    def write(self, str):
        self.text_space.insert(END, str)
        self.text_space.insert(END, '\n')
        self.text_space.see(END)
        self.text_space.update()

    def restoreStd(self):
        # 恢复标准输出
        sys.stdout = self.stdoutbak
        sys.stderr = self.stderrbak

    def flush(self):
        # 关闭程序时会调用flush刷新缓冲区，没有该函数关闭时会报错
        pass

class tkframe(tkinter.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.pack()

        # 创建数据实例对象
        self.global_data = GlobalData()

        # 创建功能实例对象
        self.message = Message(self.global_data)
        self.group = Group(self.global_data)

        # 创建对象框架，元素按照框架排布
        self.row_01 = tkinter.Frame()
        self.row_01.pack()
        self.row_02 = tkinter.Frame()
        self.row_02.pack()
        self.row_03 = tkinter.Frame()
        self.row_03.pack()
        self.row_04 = tkinter.Frame()
        self.row_04.pack()
        self.row_05 = tkinter.Frame()
        self.row_05.pack()
        self.row_06 = tkinter.Frame()
        self.row_06.pack()
        self.row_07 = tkinter.Frame()
        self.row_07.pack()
        self.row_08 = tkinter.Frame()
        self.row_08.pack()
        self.row_09 = tkinter.Frame()
        self.row_09.pack()
        self.row_10 = tkinter.Frame()
        self.row_10.pack()
        self.row_11 = tkinter.Frame()
        self.row_11.pack()

        # 创建窗口
        self.create_widgets()

    def create_widgets(self):
        # 设置开始位置标签和输入框
        self.start_num_laber = tkinter.Label(self.row_01, width=20, text="开始位置", font=('楷体', 8), justify=LEFT)
        self.start_num_laber.pack(side=LEFT)
        self.start_num = tkinter.StringVar()
        self.start_num_entry = tkinter.Entry(self.row_01, borderwidth=5, textvariable=self.start_num, font=('楷体', 8),
                                             justify=LEFT)
        self.start_num_entry.pack(side=LEFT)
        self.start_num.set(self.global_data.data["start_num"])

        # 设置聊天列表位置标签和输入框
        self.messagelist_para_laber = tkinter.Label(self.row_02, width=20, text="聊天列表", font=('楷体', 8), justify=LEFT)
        self.messagelist_para_laber.pack(side=LEFT)
        self.messagelist_para = tkinter.StringVar()
        self.messagelist_para_entry = tkinter.Entry(self.row_02, borderwidth=5, textvariable=self.messagelist_para,
                                                    font=('楷体', 8), justify=LEFT)
        self.messagelist_para_entry.pack(side=LEFT)
        self.messagelist_para.set(self.global_data.data["messagelist_para"])

        # 设置联系人/群位置标签和输入框
        self.contacts_para_laber = tkinter.Label(self.row_03, width=20, text="联系人or群", font=('楷体', 8), justify=LEFT)
        self.contacts_para_laber.pack(side=LEFT)
        self.contacts_para = tkinter.StringVar()
        self.contacts_para_entry = tkinter.Entry(self.row_03, borderwidth=5, textvariable=self.contacts_para,
                                                 font=('楷体', 8), justify=LEFT)
        self.contacts_para_entry.pack(side=LEFT)
        self.contacts_para.set(self.global_data.data["contacts_para"])

        # 设置文本偏移量
        self.text_para_laber = tkinter.Label(self.row_04, width=20, text="文本输入框", font=('楷体', 8), justify=LEFT)
        self.text_para_laber.pack(side=LEFT)
        self.text_para = tkinter.StringVar()
        self.text_para_entry = tkinter.Entry(self.row_04, borderwidth=5, textvariable=self.text_para, font=('楷体', 8),
                                             justify=LEFT)
        self.text_para_entry.pack(side=LEFT)
        self.text_para.set(self.global_data.data["contacts_para"])

        # 联系人搜索时间
        self.search_wait_time_laber = tkinter.Label(self.row_05, width=20, text="联系人搜索时间", font=('楷体', 8), justify=LEFT)
        self.search_wait_time_laber.pack(side=LEFT)
        self.search_wait_time = tkinter.DoubleVar()
        self.search_wait_time_entry = tkinter.Entry(self.row_05, borderwidth=5, textvariable=self.search_wait_time,
                                                    font=('楷体', 8), justify=LEFT)
        self.search_wait_time_entry.pack(side=LEFT)
        self.search_wait_time.set(self.global_data.data["search_wait_time"])

        # 消息间隔时间
        self.message_wait_time_laber = tkinter.Label(self.row_06, width=20, text="消息间隔时间", font=('楷体', 8), justify=LEFT)
        self.message_wait_time_laber.pack(side=LEFT)
        self.message_wait_time = tkinter.DoubleVar()
        self.message_wait_time_entry = tkinter.Entry(self.row_06, borderwidth=5, textvariable=self.message_wait_time,
                                                     font=('楷体', 8), justify=LEFT)
        self.message_wait_time_entry.pack(side=LEFT)
        self.message_wait_time.set(self.global_data.data["message_wait_time"])

        # 创建按钮并调用选择地址功能
        self.update_button = tkinter.Button(self.row_10, text="选择文件", font=('楷体', 8), command=self.read_file)
        self.update_button.pack(side=LEFT, padx=5, pady=5)

        # 创建按钮并调用更新参数功能
        self.read_file_button = tkinter.Button(self.row_10, text="更新参数", font=('楷体', 8),command=self.update_button_callback)
        self.read_file_button.pack(side=LEFT, padx=5, pady=5)

        # 创建按钮并调用更新参数功能
        self.read_file_button = tkinter.Button(self.row_10, text="发反馈", font=('楷体', 8),command=self.message.message_start)
        self.read_file_button.pack(side=LEFT, padx=5, pady=5)

        # 创建按钮并调用更新参数功能
        self.read_file_button = tkinter.Button(self.row_10, text="拉群", font=('楷体', 8),command=self.group.group_start)
        self.read_file_button.pack(side=LEFT, padx=5, pady=5)

        # 运行信息输出文本框
        self.info_text1 = tkinter.Text(self.row_11, relief="solid", width=55, height=27)  # 333*355，385
        self.info_text1.pack(pady=10)
        sys.stdout = StdoutRedirector(self.info_text1)


    def update_button_callback(self):
        self.update_global_data()

    def update_global_data(self):
        try:
            new_start_num = int(self.start_num.get())
            new_messagelist_para = int(self.messagelist_para.get())
            new_contacts_para = int(self.contacts_para.get())
            new_text_para = int(self.text_para.get())
            new_search_wait_time = double(self.search_wait_time.get())
            new_message_wait_time = double(self.message_wait_time.get())

            self.global_data.update_data({'start_num': new_start_num,'messagelist_para':new_messagelist_para,'contacts_para':new_contacts_para,'text_para':new_text_para,'search_wait_time':new_search_wait_time,'message_wait_time':new_message_wait_time})

            messagebox.showinfo("保存成功", "参数已保存！")
        except ValueError:
            self.start_num.set(str(self.global_data.data.get("start_num", 1)))
            messagebox.showerror("错误", "必须输入数字！")

    def read_file(self):
        self.filepath = filedialog.askopenfilename(title="上传文件", initialdir="f",
                                                   filetypes=[("*", "*.xlsx"), ("*", "*.xls")])
        self.global_data.update_data({"path": self.filepath})
        # print(self.global_data.data.get("path"))

    #
    #
    # def out_text(self):
    #     #运行信息输出文本框
    #     self.info_text1 = tkinter.Text(self.out_text_frame, relief="solid", width=55, height=27)#333*355，385
    #     self.info_text1.pack(pady=10)
    #     sys.stdout = StdoutRedirector(self.info_text1)
    #
    # def quit(self):
    #     self.btnQuit = Button(self.frame2, text="退出",relief="ridge", width=10, command=root.destroy, font=('微软雅黑', 8, 'bold'))
    #     self.btnQuit.pack(side=LEFT,padx=5,pady=5)



def read_excel_file(excel_file_path):
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(filename=excel_file_path)

    # 获取第一个工作表
    worksheet = workbook.active

    # 获取工作表的行列数范围
    max_row, min_row, max_col, min_col = worksheet.max_row, worksheet.min_row, worksheet.max_column, worksheet.min_column

    # 遍历工作表中的所有行和列，并将值存储到一个列表中
    data = []
    for row in range(min_row, max_row + 1):
        row_data = []
        for col in range(min_col, max_col + 1):
            # 读取单元格的值
            value = worksheet.cell(row=row, column=col).value
            if value is None:
                # 如果单元格的值为空，则将其设置为空字符串
                value = ""
            else:
                # 否则，将单元格的值转换为字符串类型
                value = str(value)
            # 将单元格的值添加到当前行数据中
            row_data.append(value)
        # 将当前行数据添加到所有数据列表中
        data.append(row_data)

    # 关闭 Excel 文件
    workbook.close()

    # 返回所有数据的列表
    return data

def check_none(item, search_wait_time):
    if item is None or item == '':
        return True
    else:
        message_search_delete = ".\\image\\message_search_delete.png"
        result = check_image_existence(message_search_delete)
        if result is None:
            temp_image = ".\\image\\message_search.png"
            x, y = find_image_location(temp_image)
            click_center(x, y, 0.2)
            pyperclip.copy(item)
            pyautogui.hotkey("ctrl", "v")
            time.sleep(search_wait_time)
        else:
            x, y = result
            click_center(x, y, 0.2)
            pyperclip.copy(item)
            pyautogui.hotkey("ctrl", "v")
            time.sleep(search_wait_time)

        return False

def get_cell_type(cell_value):
    if cell_value is None or cell_value == '':  # 判断是否为空值
        return None
    elif isinstance(cell_value, str):  # 判断是否为文本
        if os.path.isfile(cell_value):
            file_extension = os.path.splitext(cell_value)[1].lower()  # 获取文件扩展名，并转换为小写字母
            if file_extension in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:  # 判断是否为图片格式
                return 'image'
            else:
                return 'file'
        return 'text'  # 不是文件路径则返回文本类型

def process_cell_value(cell_type, input_string, message_wait_time, text_para):
    if cell_type == 'text':  # 输出文本
        temp_image = ".\\image\\image.png"
        x, y = find_image_location(temp_image)
        click_below(x, y, text_para, 0.2)

        pyperclip.copy(input_string)
        pyperclip.copy(input_string)
        pyautogui.hotkey("ctrl", "v")
        time.sleep(0.1)

        # 准备发送
        pyautogui.press("enter")
        pyautogui.press("enter")
        time.sleep(message_wait_time)

    elif cell_type == 'file':  # 输出文件
        temp_image = ".\\image\\file.png"
        x, y = find_image_location(temp_image)
        click_center(x, y, 0.2)
        pyautogui.move(0, text_para)

        # 复制粘贴文件地址
        pyperclip.copy(input_string)
        pyperclip.copy(input_string)
        pyautogui.hotkey("ctrl", "v")
        time.sleep(0.1)

        # 准备发送
        pyautogui.press("enter")
        time.sleep(0.1)
        pyautogui.press("enter")
        time.sleep(message_wait_time)

    elif cell_type == 'image':  # 输出图片
        temp_image = ".\\image\\image.png"
        x, y = find_image_location(temp_image)
        click_center(x, y, 0.2)
        pyautogui.move(0, text_para)

        # 复制粘贴图片地址
        pyperclip.copy(input_string)
        pyperclip.copy(input_string)
        pyautogui.hotkey("ctrl", "v")
        time.sleep(0.1)

        # 准备发送
        pyautogui.press("enter")
        time.sleep(0.1)
        pyautogui.press("enter")
        time.sleep(message_wait_time)

    elif cell_type is None:
        pass  # 空值不做任何处理，直接进入下一次循环

def check_image_existence(temp_image):
    # 读取搜索框
    template = cv.imread(temp_image)

    # 尝试3次查找按钮
    for i in range(2):
        # 读取截图
        pyautogui.screenshot(".\\image\\screen.png")
        source = cv.imread(".\\image\\screen.png")

        # 匹配图像
        result = cv.matchTemplate(source, template, cv.TM_CCOEFF_NORMED)

        # 找到最佳匹配位置
        min_val, max_val, min_loc, max_loc = cv.minMaxLoc(result)

        print(max_val)
        # 如果相似度大于等于0.8，则返回中心坐标
        if max_val >= 0.9:
            center_x = max_loc[0] + template.shape[1] // 2
            center_y = max_loc[1] + template.shape[0] // 2
            return center_x, center_y

    # 如果尝试了三次都没有找到，则返回None
    return None

def find_image_location(temp_image):
    """
    在源图像中查找模板图像的位置

    :param src_path: 源图像路径
    :param temp_image: 要查找的模板图像路径
    :param max_depth: 递归匹配深度，默认为10
    :param threshold: 匹配得分阈值，默认为0.9
    :return: 二元组(x, y)，表示找到的模板图像在源图像中的中心位置坐标；如果未找到匹配的位置，则返回None.
    """

    max_depth = 10
    threshold = 0.8

    # 创建源图像
    pyautogui.screenshot(".\\image\\screen.png")
    # 读取源图像
    src_path = ".\\image\\screen.png"
    source = cv.imread(src_path)

    # 读取模板图像
    template = cv.imread(temp_image)

    # 计算相似度得分矩阵
    result = cv.matchTemplate(source, template, cv.TM_CCOEFF_NORMED)

    # 找到最大得分的位置
    min_val, max_val, min_loc, max_loc = cv.minMaxLoc(result)
    x = int(max_loc[0] + template.shape[1] / 2)
    y = int(max_loc[1] + template.shape[0] / 2)

    # 根据阈值判断是否找到匹配位置
    if max_val < threshold:
        if max_depth > 1:
            # 如果未达到最大深度，则继续尝试匹配
            return find_image_location(temp_image)
        else:
            # 达到最大深度仍未匹配成功，返回None
            return None
    else:
        # 匹配成功，返回匹配位置
        return x, y

def click_center(x, y, temp_time):
    pyautogui.click(x, y, duration=0.2)
    time.sleep(temp_time)

def click_below(x, y, offset, temp_time):
    pyautogui.click(x, y + offset, duration=0.2)
    time.sleep(temp_time)


# 按间距中的绿色按钮以运行脚本。
if __name__ == '__main__':
    # message = Message()
    # message.message_start()
    # group = Group()
    # group.group_start()
    root = tkinter.Tk()
    root.title("GZ-XDF-OMO")
    # 窗口标题
    root.resizable(1, 1)
    # 窗口大小可调性
    root.geometry('400x800+100+30')
    # 窗口长宽，上下距离
    app = tkframe(master=root)
    # 循环
    root.mainloop()
